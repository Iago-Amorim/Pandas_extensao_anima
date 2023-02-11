import pandas as pd

combustiveis_df = pd.read_excel("ca-2021-02.xlsx")

print(combustiveis_df)

# Adicionar nova coluna
combustiveis_df['Ativo'] = True
print(combustiveis_df.head())

# Nova coluna com condição
combustiveis_df["Obs"] = ["MELHOR CIDADE" if municipio == "SALVADOR" else "" for municipio in combustiveis_df['Municipio']]
print(combustiveis_df.loc[combustiveis_df['Municipio'].isin(["SALVADOR", "JEQUIE"])])


import numpy as np

# Nova coluna com condição com numpy
combustiveis_df['Status Valor de Venda'] = np.where(combustiveis_df["Valor de Venda"] >= 6.5, "CARO", "BARATO")
print(combustiveis_df[['Revenda', "Valor de Venda", 'Status Valor de Venda']])

# Nova tabela 
num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv")
print(num_habitantes_df)

# Trocar nome da coluna
num_habitantes_df.rename(columns={'Estado': 'Estado - Sigla'}, inplace=True)
print(num_habitantes_df)

# Criar nova mesclando duas tabelas
colunas = ['Municipio', 'Estado - Sigla']
# merge_df = combustiveis_df.merge(num_habitantes_df, how="right", on=colunas)
# merge_df = combustiveis_df.merge(num_habitantes_df, how="left", on=colunas)
# merge_df = combustiveis_df.merge(num_habitantes_df, how="outer", on=colunas)
merge_df = combustiveis_df.merge(num_habitantes_df, how="inner", on=colunas)
print(merge_df)
print(merge_df.shape)

# Renovendo colunas vazias
merge_df.dropna(axis='columns', inplace=True)
print(merge_df.info())

# Remover colunas
print(merge_df.columns)
columns = ['Regiao - Sigla', 'Nome da Rua', 'Numero Rua', 'Bairro', 'Cep',
       'Produto', 'Valor de Venda', 'Unidade de Medida',
       'Bandeira', 'Ativo', 'Status Valor de Venda', 'Data da Coleta', 'Obs']
merge_df.drop(labels=columns, axis=1, inplace=True)
print(merge_df.info())
print(merge_df.head(100))

# Remover linhas duplicadas
merge_df.drop_duplicates(inplace=True)
print(merge_df.head(100))

# Nova tabela juntando linhas com mesmas informações
postos_por_municipio_df = merge_df.groupby(by=['Estado - Sigla', 'Municipio', "NumHabitantes2021"]).count()
postos_por_municipio_df.drop('CNPJ da Revenda', axis=1, inplace=True)
postos_por_municipio_df.reset_index(inplace=True)
postos_por_municipio_df.rename(columns={"Revenda":"Numero de Postos"}, inplace=True)
print(postos_por_municipio_df)
print(postos_por_municipio_df.info())

# Calculo de número de Habitantes por posto
postos_por_municipio_df['Num Habitante Por Postos'] = postos_por_municipio_df['NumHabitantes2021'] / postos_por_municipio_df['Numero de Postos'] 
print(postos_por_municipio_df)

# Nova tabela
num_habitante_por_postos_por_municipio_df = postos_por_municipio_df[["Municipio", "Num Habitante Por Postos"]]
num_habitante_por_postos_por_municipio_df.sort_values('Num Habitante Por Postos', inplace=True)
print(num_habitante_por_postos_por_municipio_df)

# Biblioteca de gráfico
import matplotlib.pyplot as plt

# Mostra gráfico
plt.hist(combustiveis_df['Valor de Venda'])
plt.show()
plt.hist(combustiveis_df['Valor de Venda'].mean())
plt.show()

# Gráfico personalizado
plt.hist(combustiveis_df['Valor de Venda'])
plt.title("Preço dos combustiveis Nov/2021")
plt.xlabel("Preço (em reais)")
plt.ylabel("Quantidade de Coletas")
plt.axvline(combustiveis_df['Valor de Venda'].mean(), color="#f00", linestyle='--')
plt.show()

# Tabela com média de valor de cada combustivel
c_mean = combustiveis_df['Valor de Venda'].groupby(by=combustiveis_df['Produto']).mean()
print(c_mean)

# Ajuda no grafico
import seaborn as sns

# Personalização e formatação do gráfico 
plt.figure(figsize=(10, 5))
c_mean_grafico = c_mean.plot(
    kind="barh",
    title= "Média Preço por Combustiveis",
    xlabel="Tipo de Combustivel",
    ylabel="Preço reais/litro",
    color="red",
    alpha = 0.9,
)
c_mean_grafico.set_ylabel("Tipo de Combustivel")
c_mean_grafico.set_xlabel("Preço reais/litro")
plt.grid(axis="x")
sns.despine(right=False, bottom=True)
plt.show()

# Biblioteca para editar e formatar o excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment

print(c_mean)

# Criar um novo excel
excel = "por_litro.xlsx"
c_mean.to_excel(excel, "Sumário")

# Editar e salvar o arquivo excel
wb = load_workbook(excel)
ws = wb['Sumário']
cinzinha = PatternFill("solid", fgColor="CCCCCC")
caixas = ['A1', 'B1']
for caixa in caixas:
  ws[caixa].fill = cinzinha
# print(ws['A1'].fill.fgColor.rgb)
# print(ws['A1'].fill)
MAX_ROW = ws.max_row
num_linha = 2
while num_linha <= MAX_ROW:
  coord = "B{0}".format(num_linha)
  if ws[coord].value >= 6.5:
    ws[coord].font = Font(bold=True, color="FF0000")
  num_linha += 1
# print(ws['A1'].font)
wb.save(excel)